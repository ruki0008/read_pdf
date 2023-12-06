import openpyxl

def excel():
    wb = openpyxl.load_workbook('テスト.xlsx')
    ws = wb.worksheets[0]
    last_row = ws.max_row
    values = []

    for row in ws[ f'A2:B{last_row}']:
        data = []
        for col in row:
            data.append(col.value)
        values.append(data)
    print(values)
    return values